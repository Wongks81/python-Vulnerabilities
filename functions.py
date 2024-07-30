from openpyxl import *
from constants import RESULT_HEADERS,LAPTOP_HEADERS

def checkWSExist(wb: Workbook , ws :str):
    ''' Check if the required WorkSheet exist in the Workbook'''
    if(ws in wb.sheetnames):
        # intune sheet is available
        return True
    else:
        # No intune sheet detected
        return False

def writeResultWSHeaders(wb:Workbook, ws:str, row :int):
    wb[ws].cell(row,RESULT_HEADERS.UserName.value).value = "UserName"
    wb[ws].cell(row,RESULT_HEADERS.LaptopName.value).value= "LaptopName"
    wb[ws].cell(row,RESULT_HEADERS.Title.value).value = "Vulnerability Caught"
    wb[ws].cell(row,RESULT_HEADERS.Severity.value).value = "Severity"
    wb[ws].cell(row,RESULT_HEADERS.Solution.value).value = "Solution"
    wb[ws].cell(row,RESULT_HEADERS.Results.value).value = "Location"
    wb[ws].cell(row,RESULT_HEADERS.Country.value).value = "Country"

def extractAllKEVtoSheet(reportWB : Workbook,reportWs:str, resultWB : Workbook):
    '''
        Create a KEV sheet and copy all KEVs to it for easy matching
    '''
    ws_Row =1
    kev_Row = 1

    # Create a new KEV sheet for storing KEV vulnerabilities
    resultWB.create_sheet('KEV')
    resultWB.create_sheet('KEV_LIST')
    writeResultWSHeaders(resultWB,'KEV_LIST',1)
    
    for ws_Row in range(1, reportWB[reportWs].max_row):
        if(reportWB[reportWs].cell(ws_Row,LAPTOP_HEADERS.Vulnerability.value).fill.fgColor.index != '00000000' ):
            
            # Copy the KEV title to KEV sheet in resultWB.
            resultWB['KEV'].cell(kev_Row,1).value = reportWB[reportWs].cell(ws_Row,LAPTOP_HEADERS.Vulnerability.value).value
            kev_Row +=1