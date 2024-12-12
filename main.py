from openpyxl import *
from constants import *
from functions import *
import os
import sys


def main():
    
    changeOSPath()

    # store the workbook to an object
    intuneWB = load_workbook('Intune.xlsx')
    reportWB = load_workbook('Report.xlsx')

    # Create a new workbook to store the results
    resultWB = Workbook()

    # Check if all WS required is available
    checkAllWS(intuneWB,reportWB)

    # Create KEV sheet to extract all KEVs for comparison
    extractAllKEVtoSheet(reportWB,'Laptops',resultWB)

    # Write the username to laptop results to resultWB
    writeToResultXLS(reportWB,intuneWB,resultWB)

    
    resultWB.save("result.xlsx")
    closeWB(resultWB,reportWB,intuneWB)
    

def checkAllWS(intuneWB : Workbook, reportWB : Workbook):
    ''' Check on all worksheets needed for the report'''
    if(checkWSExist(intuneWB,'Intune') == False):
        sys.exit(ERR_MSG.Intune_ws_err.value)
    
    if(checkWSExist(reportWB,REPORT_SHEET_NAME) == False):
        sys.exit(ERR_MSG.Main_ws_err.value)
    
    if(checkWSExist(reportWB,'Laptops') == False):
        sys.exit(ERR_MSG.Laptop_ws_err.value)  
    
def changeOSPath():
    # Change the folder path to the current folder 
    abspath = os.path.abspath(__file__)
    dname = os.path.dirname(os.path.abspath(sys.argv[0]))
    PROGRAM_DIR = os.path.dirname(os.path.abspath(sys.argv[0]))

    print("PATH : " + abspath)
    print("Directory : " + dname)
    print("Prog Dir : " + PROGRAM_DIR)

    os.chdir(dname)

def writeToResultXLS(reportWB: Workbook, intuneWB : Workbook,resultWB : Workbook):

    # Fill Result sheet headers
    result_Row = intune_Row = list_Row = kev_Row = 1
    writeResultWSHeaders(resultWB,'Sheet',result_Row)

    for report_Row in range(1,reportWB[REPORT_SHEET_NAME].max_row):
        # convert NetBios name for current row  in report.xlsx to string for comparison
        netBios_Report = str(reportWB[REPORT_SHEET_NAME].cell(report_Row,REPORT_HEADERS.NetBios.value).value)
        
        # Check if row value belongs to LAPTOP or India VMs
        if(('LAPTOP' in netBios_Report ) or ('THM-DEV' in netBios_Report )):
            
            # After getting the Netbios / hostname of the machine, search intune.xlsx for user details
            for intune_Row in range(1, intuneWB['Intune'].max_row):
                
                # convert Netbios name for current row in Intune.xlsx to string for comparison
                netBios_Intune = str(intuneWB['Intune'].cell(intune_Row, INTUNE_HEADERS.DeviceName.value).value)
                
                if(netBios_Report  == netBios_Intune):
                    # Found Laptop Name in Intune.xlsx, retrieve username from intune.xlsx and report.xlsx
                    resultWB['Sheet'].cell(result_Row+1,RESULT_HEADERS.UserName.value).value = intuneWB['Intune'].cell(intune_Row,INTUNE_HEADERS.UserName.value).value
                    break

            # NetBios name starts with 'LAPTOP*' or 'THM*'
            resultWB['Sheet'].cell(result_Row+1,RESULT_HEADERS.Country.value).value = intuneWB['Intune'].cell(intune_Row,INTUNE_HEADERS.Category.value).value
            resultWB['Sheet'].cell(result_Row+1,RESULT_HEADERS.LaptopName.value).value = reportWB[REPORT_SHEET_NAME].cell(report_Row,REPORT_HEADERS.NetBios.value).value
            resultWB['Sheet'].cell(result_Row+1,RESULT_HEADERS.Title.value).value = reportWB[REPORT_SHEET_NAME].cell(report_Row,REPORT_HEADERS.Title.value).value
            resultWB['Sheet'].cell(result_Row+1,RESULT_HEADERS.Severity.value).value = reportWB[REPORT_SHEET_NAME].cell(report_Row,REPORT_HEADERS.Severity.value).value
            resultWB['Sheet'].cell(result_Row+1,RESULT_HEADERS.Solution.value).value = reportWB[REPORT_SHEET_NAME].cell(report_Row,REPORT_HEADERS.Solution.value).value
            resultWB['Sheet'].cell(result_Row+1,RESULT_HEADERS.Results.value).value = reportWB[REPORT_SHEET_NAME].cell(report_Row,REPORT_HEADERS.Results.value).value

            # check if current Vulnerability is a KEV
            for kev_Row in range(1,resultWB['KEV'].max_row):

                # Vulnerability title that is in 'sheet' Sheet
                val = resultWB['Sheet'].cell(result_Row+1,RESULT_HEADERS.Title.value).value
                
                # Vulnerability title that is in the KEV sheet
                result_Value = resultWB['KEV'].cell(kev_Row,1).value

                if(val == result_Value):
                    # Found Vulnerability in 'KEV' Sheet, copy over to 'KEV_LIST'
                    resultWB['KEV_LIST'].cell(list_Row+1,RESULT_HEADERS.Country.value).value = resultWB['Sheet'].cell(result_Row +1,RESULT_HEADERS.Country.value).value
                    resultWB['KEV_LIST'].cell(list_Row+1,RESULT_HEADERS.UserName.value).value = resultWB['Sheet'].cell(result_Row +1,RESULT_HEADERS.UserName.value).value
                    resultWB['KEV_LIST'].cell(list_Row+1,RESULT_HEADERS.LaptopName.value).value = resultWB['Sheet'].cell(result_Row +1,RESULT_HEADERS.LaptopName.value).value
                    resultWB['KEV_LIST'].cell(list_Row+1,RESULT_HEADERS.Title.value).value = resultWB['Sheet'].cell(result_Row +1,RESULT_HEADERS.Title.value).value
                    resultWB['KEV_LIST'].cell(list_Row+1,RESULT_HEADERS.Severity.value).value = resultWB['Sheet'].cell(result_Row +1,RESULT_HEADERS.Severity.value).value
                    resultWB['KEV_LIST'].cell(list_Row+1,RESULT_HEADERS.Solution.value).value = resultWB['Sheet'].cell(result_Row +1,RESULT_HEADERS.Solution.value).value
                    resultWB['KEV_LIST'].cell(list_Row+1,RESULT_HEADERS.Results.value).value = resultWB['Sheet'].cell(result_Row +1,RESULT_HEADERS.Results.value).value
                    list_Row +=1
                    # Once found, do not need to loop any more
                    break

            result_Row +=1
          

                  

def closeWB(*args :Workbook):
    for wb in args:
        wb.close()

if __name__ == "__main__":
    main()